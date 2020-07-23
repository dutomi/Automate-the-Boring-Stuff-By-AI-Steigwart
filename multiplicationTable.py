#! python3
# This program takes a value N in the command line
# and creates an NxN multiplication table in an Excel spreadsheet


import openpyxl
import sys
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']

#BoldFont
BoldFont = Font(bold = True)

#Setting the elements in column A
for i in range(2,int(sys.argv[1])+2):
    sheet['A' + str(i)] = i-1   # sets the elements on the column A from 1 to the value in sys.argv[1]
    sheet['A' + str(i)].font = BoldFont



# setting the elements in row 1
columns = 2
while columns <= int(sys.argv[1])+1:
    sheet.cell(row = 1,column=columns).value = columns - 1
    sheet.cell(row=1, column=columns).font = BoldFont
    columns += 1


#setting the formulas inside the cells for actual multiplication
rows = 2
while rows <= sheet.max_row:
    for columnnum in range(2,sheet.max_column+1):
        sheet.cell(row=rows,column=columnnum).value = '={}'.format(sheet.cell(row=rows,column=1).value * sheet.cell(row=1,column=columnnum).value)
    rows += 1

wb.save('multiplicationtablecopy.xlsx')

